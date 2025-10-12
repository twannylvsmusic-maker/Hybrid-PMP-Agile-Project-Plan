#!/usr/bin/env python3
"""
update_weekly.py - Enhanced Version
Ingest weekly CSVs and refresh the "Weekly_Updates" sheet in Hybrid_ProjectPlan_Template.xlsx.

CSV schema (header row required):
Week Start, Tickets Opened, Tickets Resolved, NPS, CSAT, Notes

Features:
- Enhanced error handling and validation
- Email notifications on errors
- Data quality checks
- Automatic trend analysis
- Logging capabilities
- Configuration file support

Usage:
  1) Set the config values below (paths).
  2) Run: python update_weekly.py
  3) Optional: Schedule weekly via Windows Task Scheduler or cron.

Requirements:
  pip install openpyxl pandas python-dateutil
"""

import os
import glob
import logging
import json
import smtplib
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# ------------------
# Enhanced Config
# ------------------
CONFIG = {
    "EXCEL_PATH": r"../Hybrid_ProjectPlan_Template.xlsx",
    "WEEKLY_SHEET": "Weekly_Updates",
    "CSV_FOLDER": r"./weekly_csvs",
    "START_ROW": 4,
    "BACKUP_ON_SAVE": True,
    "LOG_LEVEL": "INFO",
    "LOG_FILE": "weekly_update.log",
    "EMAIL_NOTIFICATIONS": {
        "ENABLED": False,
        "SMTP_SERVER": "smtp.gmail.com",
        "SMTP_PORT": 587,
        "EMAIL": "",
        "PASSWORD": "",
        "TO_EMAILS": []
    },
    "DATA_VALIDATION": {
        "MIN_CSAT": 0,
        "MAX_CSAT": 10,
        "MIN_NPS": -100,
        "MAX_NPS": 100,
        "MAX_TICKETS": 10000
    }
}

# Load config from file if exists
CONFIG_FILE = "config.json"
if os.path.exists(CONFIG_FILE):
    with open(CONFIG_FILE, 'r') as f:
        user_config = json.load(f)
        CONFIG.update(user_config)

# ------------------
# Logging Setup
# ------------------
def setup_logging():
    logging.basicConfig(
        level=getattr(logging, CONFIG["LOG_LEVEL"]),
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(CONFIG["LOG_FILE"]),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ------------------
# Email Notifications
# ------------------
def send_email_notification(subject: str, body: str, is_error: bool = False):
    """Send email notification if enabled"""
    if not CONFIG["EMAIL_NOTIFICATIONS"]["ENABLED"]:
        return
    
    try:
        msg = MIMEMultipart()
        msg['From'] = CONFIG["EMAIL_NOTIFICATIONS"]["EMAIL"]
        msg['To'] = ", ".join(CONFIG["EMAIL_NOTIFICATIONS"]["TO_EMAILS"])
        msg['Subject'] = f"{'[ERROR]' if is_error else '[INFO]'} {subject}"
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(CONFIG["EMAIL_NOTIFICATIONS"]["SMTP_SERVER"], 
                             CONFIG["EMAIL_NOTIFICATIONS"]["SMTP_PORT"])
        server.starttls()
        server.login(CONFIG["EMAIL_NOTIFICATIONS"]["EMAIL"], 
                    CONFIG["EMAIL_NOTIFICATIONS"]["PASSWORD"])
        server.send_message(msg)
        server.quit()
        logger.info(f"Email notification sent: {subject}")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

# ------------------
# Data Validation
# ------------------
def validate_data_quality(df: pd.DataFrame) -> List[str]:
    """Validate data quality and return list of issues"""
    issues = []
    
    # Check for missing values
    missing_data = df.isnull().sum()
    for col, count in missing_data.items():
        if count > 0:
            issues.append(f"Missing values in {col}: {count} rows")
    
    # Check data ranges
    if 'CSAT' in df.columns:
        invalid_csat = df[(df['CSAT'] < CONFIG["DATA_VALIDATION"]["MIN_CSAT"]) | 
                         (df['CSAT'] > CONFIG["DATA_VALIDATION"]["MAX_CSAT"])]
        if len(invalid_csat) > 0:
            issues.append(f"Invalid CSAT values: {len(invalid_csat)} rows outside range 0-10")
    
    if 'NPS' in df.columns:
        invalid_nps = df[(df['NPS'] < CONFIG["DATA_VALIDATION"]["MIN_NPS"]) | 
                        (df['NPS'] > CONFIG["DATA_VALIDATION"]["MAX_NPS"])]
        if len(invalid_nps) > 0:
            issues.append(f"Invalid NPS values: {len(invalid_nps)} rows outside range -100 to 100")
    
    # Check for future dates
    future_dates = df[df['Week Start'] > datetime.now()]
    if len(future_dates) > 0:
        issues.append(f"Future dates detected: {len(future_dates)} rows")
    
    return issues

# ------------------
# Trend Analysis
# ------------------
def analyze_trends(df: pd.DataFrame) -> Dict:
    """Analyze trends in the data"""
    if len(df) < 2:
        return {"error": "Insufficient data for trend analysis"}
    
    trends = {}
    
    # CSAT trend
    if 'CSAT' in df.columns and len(df[df['CSAT'].notna()]) >= 2:
        recent_csat = df['CSAT'].tail(4).mean()
        overall_csat = df['CSAT'].mean()
        trends['CSAT'] = {
            'recent_avg': round(recent_csat, 2),
            'overall_avg': round(overall_csat, 2),
            'trend': 'improving' if recent_csat > overall_csat else 'declining'
        }
    
    # Ticket resolution trend
    if 'Tickets Resolved' in df.columns:
        recent_resolved = df['Tickets Resolved'].tail(4).sum()
        overall_resolved = df['Tickets Resolved'].sum()
        trends['Resolution'] = {
            'recent_total': int(recent_resolved),
            'trend': 'improving' if recent_resolved > overall_resolved * 0.25 else 'stable'
        }
    
    return trends

def read_all_weekly_csvs(csv_folder: str) -> pd.DataFrame:
    """Enhanced CSV reading with better error handling"""
    files = sorted(glob.glob(os.path.join(csv_folder, "*.csv")))
    frames = []
    processed_files = []
    
    if not files:
        logger.warning(f"No CSV files found in {csv_folder}")
        return pd.DataFrame(columns=["Week Start","Tickets Opened","Tickets Resolved","NPS","CSAT","Notes"])
    
    for f in files:
        try:
            df = pd.read_csv(f)
            if df.empty:
                logger.warning(f"Empty file skipped: {f}")
                continue
            frames.append(df)
            processed_files.append(f)
            logger.info(f"Successfully processed: {f}")
        except Exception as e:
            error_msg = f"Failed to process {f}: {e}"
            logger.error(error_msg)
            send_email_notification("CSV Processing Error", error_msg, is_error=True)
    
    if not frames:
        raise ValueError("No valid CSV files could be processed")
    
    df = pd.concat(frames, ignore_index=True)
    logger.info(f"Combined data from {len(processed_files)} files: {len(df)} rows")
    
    # Normalize columns
    expected = ["Week Start","Tickets Opened","Tickets Resolved","NPS","CSAT","Notes"]
    df = df.rename(columns={c:c.strip() for c in df.columns})
    missing = [c for c in expected if c not in df.columns]
    if missing:
        error_msg = f"Missing required columns: {missing}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Type conversions with better error handling
    df["Week Start"] = pd.to_datetime(df["Week Start"], errors="coerce")
    for col in ["Tickets Opened","Tickets Resolved","NPS","CSAT"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # Data quality validation
    quality_issues = validate_data_quality(df)
    if quality_issues:
        logger.warning(f"Data quality issues found: {'; '.join(quality_issues)}")
        if CONFIG["EMAIL_NOTIFICATIONS"]["ENABLED"]:
            send_email_notification("Data Quality Issues", 
                                  "Issues found:\n" + "\n".join(quality_issues))
    
    # Deduplicate by Week Start
    initial_count = len(df)
    df = df.sort_values("Week Start").drop_duplicates(subset=["Week Start"], keep="last")
    final_count = len(df)
    if initial_count != final_count:
        logger.info(f"Removed {initial_count - final_count} duplicate entries")
    
    return df

def write_to_excel(excel_path: str, sheet: str, df: pd.DataFrame, start_row: int = 4):
    """Enhanced Excel writing with better error handling"""
    try:
        # Validate Excel file exists
        if not os.path.exists(excel_path):
            error_msg = f"Excel file not found: {excel_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        # Load workbook with error handling
        try:
            wb = load_workbook(excel_path)
        except InvalidFileException as e:
            error_msg = f"Invalid Excel file: {e}"
            logger.error(error_msg)
            raise
        
        # Validate sheet exists
        if sheet not in wb.sheetnames:
            error_msg = f"Sheet '{sheet}' not found in Excel file"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        ws = wb[sheet]
        
        # Clear old data region (columns A:F for safety)
        max_row = ws.max_row
        for r in range(start_row, max_row+1):
            for c in range(1, 7):
                ws.cell(row=r, column=c, value=None)
        
        # Write new rows with data validation
        df = df.sort_values("Week Start")
        r = start_row
        written_rows = 0
        
        for _, row in df.iterrows():
            try:
                ws.cell(row=r, column=1, value=row["Week Start"].date() if pd.notnull(row["Week Start"]) else None)
                ws.cell(row=r, column=2, value=int(row["Tickets Opened"]) if pd.notnull(row["Tickets Opened"]) else None)
                ws.cell(row=r, column=3, value=int(row["Tickets Resolved"]) if pd.notnull(row["Tickets Resolved"]) else None)
                ws.cell(row=r, column=4, value=float(row["NPS"]) if pd.notnull(row["NPS"]) else None)
                ws.cell(row=r, column=5, value=float(row["CSAT"]) if pd.notnull(row["CSAT"]) else None)
                ws.cell(row=r, column=6, value=str(row["Notes"]) if pd.notnull(row["Notes"]) else None)
                written_rows += 1
                r += 1
            except Exception as e:
                logger.error(f"Error writing row {r}: {e}")
                continue
        
        # Create backup if enabled
        if CONFIG["BACKUP_ON_SAVE"]:
            root, ext = os.path.splitext(excel_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{root}_backup_{timestamp}{ext}"
            import shutil
            shutil.copy2(excel_path, backup_path)
            logger.info(f"Backup created: {backup_path}")
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Successfully wrote {written_rows} rows into {sheet} @ {excel_path}")
        
        return written_rows
        
    except Exception as e:
        error_msg = f"Failed to write to Excel: {e}"
        logger.error(error_msg)
        send_email_notification("Excel Write Error", error_msg, is_error=True)
        raise

# ------------------
# Configuration Management
# ------------------
def create_default_config():
    """Create a default configuration file"""
    config_path = "config.json"
    if not os.path.exists(config_path):
        with open(config_path, 'w') as f:
            json.dump(CONFIG, f, indent=2)
        logger.info(f"Created default config file: {config_path}")
        print(f"Created default config file: {config_path}")
        print("Please edit config.json to customize settings")

# ------------------
# Main Execution
# ------------------
def main():
    """Enhanced main function with comprehensive error handling"""
    start_time = datetime.now()
    logger.info("Starting weekly update process")
    
    try:
        # Create default config if needed
        create_default_config()
        
        # Ensure CSV folder exists
        os.makedirs(CONFIG["CSV_FOLDER"], exist_ok=True)
        
        # Read and process CSV files
        df = read_all_weekly_csvs(CONFIG["CSV_FOLDER"])
        if df.empty:
            logger.info("No CSVs found; nothing to update.")
            send_email_notification("Weekly Update - No Data", 
                                  "No CSV files found for processing this week.")
            return
        
        # Perform trend analysis
        trends = analyze_trends(df)
        if trends:
            logger.info(f"Trend analysis: {trends}")
        
        # Write to Excel
        written_rows = write_to_excel(CONFIG["EXCEL_PATH"], CONFIG["WEEKLY_SHEET"], df, CONFIG["START_ROW"])
        
        # Success notification
        duration = datetime.now() - start_time
        success_msg = f"Weekly update completed successfully!\n\n"
        success_msg += f"• Processed {len(df)} data rows\n"
        success_msg += f"• Written {written_rows} rows to Excel\n"
        success_msg += f"• Duration: {duration.total_seconds():.2f} seconds\n"
        
        if trends and 'error' not in trends:
            success_msg += f"• Trends detected: {trends}\n"
        
        logger.info(success_msg.replace('\n', ' '))
        
        if CONFIG["EMAIL_NOTIFICATIONS"]["ENABLED"]:
            send_email_notification("Weekly Update Success", success_msg)
        
    except Exception as e:
        error_msg = f"Weekly update failed: {e}"
        logger.error(error_msg)
        send_email_notification("Weekly Update Failed", error_msg, is_error=True)
        raise

if __name__ == "__main__":
    main()
