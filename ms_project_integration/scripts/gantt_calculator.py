"""
Gantt Chart Calculator
Handles timeline calculations, dependency resolution, and critical path analysis
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import logging
import json
import os
import argparse
from collections import defaultdict

# Configuration loading
def load_config():
    """Load configuration from JSON file"""
    config_path = os.path.join(os.path.dirname(__file__), 'config_msproject.json')
    with open(config_path, 'r') as f:
        return json.load(f)

CONFIG = load_config()

# Logging setup
def setup_logging():
    """Set up logging configuration"""
    log_level = getattr(logging, CONFIG['logging']['level'])
    log_file = CONFIG['logging']['file']
    
    # Create logs directory if it doesn't exist
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    
    handlers = [logging.FileHandler(log_file)]
    if CONFIG['logging']['console']:
        handlers.append(logging.StreamHandler())
    
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )

setup_logging()

class GanttCalculator:
    """Calculates timelines, dependencies, and critical path for Gantt charts"""
    
    def __init__(self, excel_file_path, sheet_name='Gantt Chart'):
        self.excel_file_path = excel_file_path
        self.sheet_name = sheet_name
        self.tasks = {}
        self.task_order = []
        self.working_days = self._get_working_days()
        
    def _get_working_days(self):
        """Get list of working days from config"""
        days = []
        day_map = {
            'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
            'friday': 4, 'saturday': 5, 'sunday': 6
        }
        for day, is_working in CONFIG['working_days'].items():
            if is_working:
                days.append(day_map[day])
        return days
    
    def load_tasks(self):
        """Load tasks from Excel"""
        try:
            logging.info(f"Loading tasks from: {self.excel_file_path}")
            
            wb = openpyxl.load_workbook(self.excel_file_path)
            if self.sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{self.sheet_name}' not found")
                return False
            
            ws = wb[self.sheet_name]
            
            # Read headers
            headers = [cell.value for cell in ws[1]]
            
            # Read tasks
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check Task ID
                    task_data = dict(zip(headers, row))
                    task_id = task_data.get('Task ID')
                    self.tasks[task_id] = task_data
                    self.task_order.append(task_id)
            
            logging.info(f"Loaded {len(self.tasks)} tasks")
            return True
            
        except Exception as e:
            logging.error(f"Error loading tasks: {e}")
            return False
    
    def calculate_dates(self):
        """Calculate start and finish dates based on dependencies"""
        logging.info("Calculating task dates...")
        
        calculated = {}
        
        for task_id in self.task_order:
            task = self.tasks[task_id]
            
            # If start date is already set and no dependencies, use it
            start_date = task.get('Start Date')
            duration = task.get('Duration (Days)', 0)
            dependencies = task.get('Dependencies', '')
            
            # Calculate start date based on dependencies
            if dependencies:
                dep_finish_date = self._get_latest_dependency_finish(dependencies, calculated)
                if dep_finish_date:
                    start_date = self._add_working_days(dep_finish_date, 1)
            
            # If no start date, use today
            if not start_date or not isinstance(start_date, datetime):
                start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            
            # Calculate finish date
            if duration:
                finish_date = self._add_working_days(start_date, int(float(duration)))
            else:
                finish_date = start_date
            
            calculated[task_id] = {
                'Start Date': start_date,
                'Finish Date': finish_date,
                'Duration (Days)': duration
            }
        
        return calculated
    
    def _get_latest_dependency_finish(self, dependencies, calculated):
        """Get the latest finish date from dependencies"""
        dep_finish_dates = []
        
        # Parse dependencies (e.g., "2FS", "3SS+2", "4,5")
        dep_ids = []
        for dep in str(dependencies).split(','):
            dep = dep.strip()
            if dep:
                # Extract numeric ID
                dep_id = ''.join(filter(str.isdigit, dep))
                if dep_id:
                    dep_ids.append(int(dep_id) if dep_id.isdigit() else dep_id)
        
        # Get finish dates
        for dep_id in dep_ids:
            if dep_id in calculated:
                dep_finish_dates.append(calculated[dep_id]['Finish Date'])
        
        return max(dep_finish_dates) if dep_finish_dates else None
    
    def _add_working_days(self, start_date, days):
        """Add working days to a date (skipping weekends)"""
        current_date = start_date
        days_added = 0
        
        while days_added < days:
            current_date += timedelta(days=1)
            if current_date.weekday() in self.working_days:
                days_added += 1
        
        return current_date
    
    def calculate_critical_path(self):
        """Calculate the critical path through the project"""
        logging.info("Calculating critical path...")
        
        # Calculate earliest start/finish times (forward pass)
        earliest_times = self._forward_pass()
        
        # Calculate latest start/finish times (backward pass)
        latest_times = self._backward_pass(earliest_times)
        
        # Identify critical tasks (where slack = 0)
        critical_tasks = []
        for task_id in self.task_order:
            earliest_finish = earliest_times[task_id]['finish']
            latest_finish = latest_times[task_id]['finish']
            slack = (latest_finish - earliest_finish).days
            
            if slack == 0:
                critical_tasks.append(task_id)
        
        logging.info(f"Critical path contains {len(critical_tasks)} tasks")
        return critical_tasks
    
    def _forward_pass(self):
        """Forward pass: calculate earliest start and finish times"""
        earliest = {}
        
        for task_id in self.task_order:
            task = self.tasks[task_id]
            start_date = task.get('Start Date', datetime.now())
            duration = task.get('Duration (Days)', 0)
            dependencies = task.get('Dependencies', '')
            
            # Calculate earliest start based on dependencies
            if dependencies:
                dep_finish_dates = []
                for dep in str(dependencies).split(','):
                    dep_id = ''.join(filter(str.isdigit, dep.strip()))
                    if dep_id and int(dep_id) in earliest:
                        dep_finish_dates.append(earliest[int(dep_id)]['finish'])
                
                if dep_finish_dates:
                    earliest_start = max(dep_finish_dates) + timedelta(days=1)
                else:
                    earliest_start = start_date
            else:
                earliest_start = start_date
            
            earliest_finish = self._add_working_days(earliest_start, int(float(duration)) if duration else 0)
            
            earliest[task_id] = {
                'start': earliest_start,
                'finish': earliest_finish
            }
        
        return earliest
    
    def _backward_pass(self, earliest_times):
        """Backward pass: calculate latest start and finish times"""
        latest = {}
        
        # Start from the last task
        last_task_id = self.task_order[-1]
        latest[last_task_id] = {
            'finish': earliest_times[last_task_id]['finish'],
            'start': earliest_times[last_task_id]['start']
        }
        
        # Work backwards
        for task_id in reversed(self.task_order[:-1]):
            task = self.tasks[task_id]
            duration = task.get('Duration (Days)', 0)
            
            # Find successor tasks
            successors = self._find_successors(task_id)
            
            if successors:
                latest_finish = min([latest[succ]['start'] for succ in successors if succ in latest]) - timedelta(days=1)
            else:
                latest_finish = earliest_times[task_id]['finish']
            
            latest_start = latest_finish - timedelta(days=int(float(duration)) if duration else 0)
            
            latest[task_id] = {
                'start': latest_start,
                'finish': latest_finish
            }
        
        return latest
    
    def _find_successors(self, task_id):
        """Find tasks that depend on the given task"""
        successors = []
        
        for tid, task in self.tasks.items():
            dependencies = str(task.get('Dependencies', ''))
            if str(task_id) in dependencies:
                successors.append(tid)
        
        return successors
    
    def update_excel(self, calculated_dates, critical_tasks):
        """Update Excel with calculated dates and critical path markers"""
        try:
            logging.info("Updating Excel with calculated values...")
            
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb[self.sheet_name]
            
            # Find column indices
            headers = {cell.value: cell.column for cell in ws[1]}
            
            # Update each task
            for row_idx, task_id in enumerate(self.task_order, start=2):
                if task_id in calculated_dates:
                    calc = calculated_dates[task_id]
                    
                    # Update dates
                    if 'Start Date' in headers:
                        ws.cell(row=row_idx, column=headers['Start Date'], value=calc['Start Date'])
                    
                    if 'Finish Date' in headers:
                        ws.cell(row=row_idx, column=headers['Finish Date'], value=calc['Finish Date'])
                    
                    # Mark critical tasks
                    if 'Critical Path' in headers and task_id in critical_tasks:
                        ws.cell(row=row_idx, column=headers['Critical Path'], value='Yes')
            
            wb.save(self.excel_file_path)
            logging.info("Excel updated successfully")
            return True
            
        except Exception as e:
            logging.error(f"Error updating Excel: {e}")
            return False

def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Calculate Gantt chart timelines and critical path')
    parser.add_argument('--input', '-i', required=True, help='Path to Excel Gantt chart file')
    parser.add_argument('--sheet', '-s', default='Gantt Chart', help='Excel sheet name (default: Gantt Chart)')
    parser.add_argument('--recalculate', '-r', action='store_true', help='Recalculate all dates based on dependencies')
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.input):
        logging.error(f"Input file not found: {args.input}")
        return
    
    # Calculate Gantt chart
    calculator = GanttCalculator(args.input, args.sheet)
    
    if calculator.load_tasks():
        if args.recalculate:
            calculated_dates = calculator.calculate_dates()
        else:
            calculated_dates = {}
        
        critical_tasks = calculator.calculate_critical_path()
        
        if calculator.update_excel(calculated_dates, critical_tasks):
            print(f"Successfully updated Gantt chart")
            print(f"   Tasks: {len(calculator.tasks)}")
            print(f"   Critical path: {len(critical_tasks)} tasks")
        else:
            print("Failed to update Excel")
    else:
        print("Failed to load tasks")

if __name__ == "__main__":
    main()

