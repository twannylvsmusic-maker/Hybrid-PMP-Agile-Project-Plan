"""
MS Project XML Exporter
Converts Excel Gantt chart format to MS Project XML files
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from lxml import etree
import logging
import json
import os
import argparse

# Configuration loading
def load_config():
    """Load configuration from JSON file"""
    config_path = os.path.join(os.path.dirname(__file__), 'config_msproject.json')
    with open(config_path, 'r') as f:
        return json.load(f)

CONFIG = load_config()

# Logging setup
def setup_logging():
    """Set up logging configuration"""
    log_level = getattr(logging, CONFIG['logging']['level'])
    log_file = CONFIG['logging']['file']
    
    # Create logs directory if it doesn't exist
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    
    handlers = [logging.FileHandler(log_file)]
    if CONFIG['logging']['console']:
        handlers.append(logging.StreamHandler())
    
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )

setup_logging()

class MSProjectExporter:
    """Exports Excel Gantt chart format to MS Project XML files"""
    
    def __init__(self, excel_file_path, sheet_name='Gantt Chart'):
        self.excel_file_path = excel_file_path
        self.sheet_name = sheet_name
        self.tasks = []
        
    def load_excel(self):
        """Load Excel Gantt chart data"""
        try:
            logging.info(f"Loading Excel file: {self.excel_file_path}")
            
            # Load workbook and worksheet
            wb = openpyxl.load_workbook(self.excel_file_path)
            if self.sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{self.sheet_name}' not found in workbook")
                return False
            
            ws = wb[self.sheet_name]
            
            # Read headers
            headers = [cell.value for cell in ws[1]]
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Skip empty rows (check Task ID)
                    task_data = dict(zip(headers, row))
                    self.tasks.append(task_data)
            
            logging.info(f"Successfully loaded {len(self.tasks)} tasks from Excel")
            return True
            
        except Exception as e:
            logging.error(f"Error loading Excel file: {e}")
            return False
    
    def to_ms_project_xml(self, output_path=None):
        """Convert tasks to MS Project XML format"""
        if not self.tasks:
            logging.error("No tasks to export")
            return False
        
        try:
            # Generate output path if not provided
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(
                    CONFIG['paths']['default_export_path'],
                    f'exported_project_{timestamp}.xml'
                )
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Build XML structure
            root = self._create_xml_structure()
            
            # Write XML to file
            tree = etree.ElementTree(root)
            tree.write(output_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            
            logging.info(f"Successfully exported to: {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error exporting to MS Project XML: {e}")
            return False
    
    def _create_xml_structure(self):
        """Create MS Project XML structure"""
        # Create root element with namespace
        ns = "http://schemas.microsoft.com/project"
        root = etree.Element(
            'Project',
            xmlns=ns,
            nsmap={None: ns}
        )
        
        # Add project metadata
        self._add_project_metadata(root)
        
        # Add tasks
        tasks_element = etree.SubElement(root, 'Tasks')
        for idx, task in enumerate(self.tasks, start=1):
            self._add_task(tasks_element, task, idx)
        
        return root
    
    def _add_project_metadata(self, root):
        """Add project-level metadata"""
        # Project name
        name = etree.SubElement(root, 'Name')
        name.text = 'Hybrid PMP Project'
        
        # Creation date
        creation_date = etree.SubElement(root, 'CreationDate')
        creation_date.text = datetime.now().isoformat()
        
        # Last saved
        last_saved = etree.SubElement(root, 'LastSaved')
        last_saved.text = datetime.now().isoformat()
        
        # Currency
        currency = etree.SubElement(root, 'CurrencySymbol')
        currency.text = CONFIG['ms_project_xml']['currency_symbol']
        
        # Default task type
        default_task_type = etree.SubElement(root, 'DefaultTaskType')
        default_task_type.text = '1'  # 0=FixedUnits, 1=FixedDuration, 2=FixedWork
    
    def _add_task(self, parent, task_data, task_id):
        """Add a single task to XML"""
        task = etree.SubElement(parent, 'Task')
        
        # Task ID
        uid = etree.SubElement(task, 'UID')
        uid.text = str(task_id)
        
        task_id_elem = etree.SubElement(task, 'ID')
        task_id_elem.text = str(task_data.get('Task ID', task_id))
        
        # Task Name
        name = etree.SubElement(task, 'Name')
        name.text = str(task_data.get('Task Name', 'Unnamed Task'))
        
        # Start and Finish dates
        start_date = task_data.get('Start Date')
        if isinstance(start_date, datetime):
            start = etree.SubElement(task, 'Start')
            start.text = start_date.isoformat()
        
        finish_date = task_data.get('Finish Date')
        if isinstance(finish_date, datetime):
            finish = etree.SubElement(task, 'Finish')
            finish.text = finish_date.isoformat()
        
        # Duration (convert days to PT format: PT40H0M0S for 5 days)
        duration_days = task_data.get('Duration (Days)', 0)
        if duration_days and duration_days > 0:
            duration = etree.SubElement(task, 'Duration')
            duration_hours = int(float(duration_days) * 8)  # 8-hour workday
            duration.text = f'PT{duration_hours}H0M0S'
        
        # Progress (PercentComplete)
        progress = task_data.get('Progress (%)', 0)
        if progress:
            percent_complete = etree.SubElement(task, 'PercentComplete')
            percent_complete.text = str(int(progress))
        
        # Priority
        priority_str = task_data.get('Priority', 'Medium')
        priority = etree.SubElement(task, 'Priority')
        priority_map = {'High': 800, 'Medium': 500, 'Low': 200}
        priority.text = str(priority_map.get(priority_str, 500))
        
        # Notes
        notes_text = task_data.get('Notes', '')
        if notes_text:
            notes = etree.SubElement(task, 'Notes')
            notes.text = str(notes_text)
        
        # Milestone
        status = task_data.get('Status', '')
        if 'milestone' in str(task_data.get('Task Name', '')).lower():
            milestone = etree.SubElement(task, 'Milestone')
            milestone.text = '1'
        
        # Critical
        critical = etree.SubElement(task, 'Critical')
        critical.text = '0'  # Will be calculated by MS Project
        
        # Predecessors (dependencies)
        dependencies = task_data.get('Dependencies', '')
        if dependencies:
            predecessor_link = etree.SubElement(task, 'PredecessorLink')
            # Parse dependency string (e.g., "2FS", "3SS+2")
            # This is a simplified version - full parsing would be more complex
            try:
                if dependencies.strip():
                    pred_id = ''.join(filter(str.isdigit, str(dependencies)))
                    if pred_id:
                        pred_uid = etree.SubElement(predecessor_link, 'PredecessorUID')
                        pred_uid.text = pred_id
                        
                        # Link type (default to Finish-to-Start)
                        link_type = etree.SubElement(predecessor_link, 'Type')
                        link_type.text = '1'  # 1=FS, 2=SS, 3=FF, 4=SF
            except:
                pass

def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Export Excel Gantt chart to MS Project XML')
    parser.add_argument('--input', '-i', required=True, help='Path to Excel Gantt chart file')
    parser.add_argument('--output', '-o', help='Path to output MS Project XML file (optional)')
    parser.add_argument('--sheet', '-s', default='Gantt Chart', help='Excel sheet name (default: Gantt Chart)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.input):
        logging.error(f"Input file not found: {args.input}")
        return
    
    # Export to MS Project
    exporter = MSProjectExporter(args.input, args.sheet)
    
    if exporter.load_excel():
        if exporter.to_ms_project_xml(args.output):
            print(f"Successfully exported {len(exporter.tasks)} tasks to MS Project XML")
        else:
            print("Failed to export to MS Project XML")
    else:
        print("Failed to load Excel file")

if __name__ == "__main__":
    main()

