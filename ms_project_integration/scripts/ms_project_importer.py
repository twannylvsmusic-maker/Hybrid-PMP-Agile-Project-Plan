"""
MS Project XML Importer
Converts MS Project XML files into Excel Gantt chart format
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from lxml import etree
import logging
import json
import os
import argparse

# Configuration loading
def load_config():
    """Load configuration from JSON file"""
    config_path = os.path.join(os.path.dirname(__file__), 'config_msproject.json')
    with open(config_path, 'r') as f:
        return json.load(f)

CONFIG = load_config()

# Logging setup
def setup_logging():
    """Set up logging configuration"""
    log_level = getattr(logging, CONFIG['logging']['level'])
    log_file = CONFIG['logging']['file']
    
    # Create logs directory if it doesn't exist
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    
    handlers = [logging.FileHandler(log_file)]
    if CONFIG['logging']['console']:
        handlers.append(logging.StreamHandler())
    
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )

setup_logging()

class MSProjectImporter:
    """Imports MS Project XML files and converts to Excel Gantt format"""
    
    def __init__(self, xml_file_path):
        self.xml_file_path = xml_file_path
        self.tasks = []
        self.namespace = {}
        
    def parse_xml(self):
        """Parse MS Project XML file"""
        try:
            logging.info(f"Parsing MS Project XML: {self.xml_file_path}")
            tree = etree.parse(self.xml_file_path)
            root = tree.getroot()
            
            # Extract namespace if present
            if root.tag.startswith('{'):
                self.namespace = {'ms': root.tag.split('}')[0].strip('{')}
            
            # Extract tasks
            self._extract_tasks(root)
            
            logging.info(f"Successfully parsed {len(self.tasks)} tasks from XML")
            return True
            
        except Exception as e:
            logging.error(f"Error parsing XML file: {e}")
            return False
    
    def _extract_tasks(self, root):
        """Extract task information from XML"""
        # Handle with or without namespace
        if self.namespace:
            tasks_xpath = './/ms:Task'
        else:
            tasks_xpath = './/Task'
        
        tasks = root.xpath(tasks_xpath, namespaces=self.namespace) if self.namespace else root.findall('.//Task')
        
        for task in tasks:
            task_data = {
                'Task_ID': self._get_text(task, 'ID', ''),
                'Task_Name': self._get_text(task, 'Name', 'Unnamed Task'),
                'Start_Date': self._parse_date(self._get_text(task, 'Start', '')),
                'Finish_Date': self._parse_date(self._get_text(task, 'Finish', '')),
                'Duration_Days': self._parse_duration(self._get_text(task, 'Duration', '0')),
                'Progress': self._parse_percent(self._get_text(task, 'PercentComplete', '0')),
                'Priority': self._parse_priority(self._get_text(task, 'Priority', '500')),
                'Assigned_To': self._get_text(task, 'ResourceNames', ''),
                'Predecessors': self._get_text(task, 'PredecessorLink', ''),
                'Notes': self._get_text(task, 'Notes', ''),
                'Milestone': self._get_text(task, 'Milestone', '0') == '1',
                'Critical': self._get_text(task, 'Critical', '0') == '1'
            }
            
            # Determine status based on progress
            if task_data['Progress'] == 0:
                task_data['Status'] = 'Not Started'
            elif task_data['Progress'] == 100:
                task_data['Status'] = 'Completed'
            else:
                task_data['Status'] = 'In Progress'
            
            self.tasks.append(task_data)
    
    def _get_text(self, element, tag, default=''):
        """Get text from XML element"""
        if self.namespace:
            child = element.find(f'ms:{tag}', namespaces=self.namespace)
        else:
            child = element.find(tag)
        return child.text if child is not None and child.text else default
    
    def _parse_date(self, date_str):
        """Parse MS Project date format"""
        if not date_str:
            return None
        try:
            # MS Project date format: 2024-10-17T00:00:00
            return datetime.fromisoformat(date_str.split('T')[0])
        except:
            return None
    
    def _parse_duration(self, duration_str):
        """Parse MS Project duration format (e.g., PT40H0M0S for 5 days)"""
        if not duration_str or duration_str == '0':
            return 0
        try:
            # Simple conversion: PT40H = 5 days (8-hour workday)
            if 'H' in duration_str:
                hours = int(duration_str.split('PT')[1].split('H')[0])
                return hours / 8  # Convert to days
            return 0
        except:
            return 0
    
    def _parse_percent(self, percent_str):
        """Parse percentage (0-100)"""
        try:
            return int(percent_str)
        except:
            return 0
    
    def _parse_priority(self, priority_str):
        """Parse MS Project priority (0-1000, 500=Medium)"""
        try:
            priority_int = int(priority_str)
            if priority_int >= 700:
                return 'High'
            elif priority_int >= 300:
                return 'Medium'
            else:
                return 'Low'
        except:
            return 'Medium'
    
    def to_dataframe(self):
        """Convert tasks to pandas DataFrame"""
        if not self.tasks:
            logging.warning("No tasks to convert")
            return pd.DataFrame()
        
        df = pd.DataFrame(self.tasks)
        logging.info(f"Created DataFrame with {len(df)} tasks")
        return df
    
    def to_excel(self, output_path=None):
        """Export tasks to Excel Gantt chart template"""
        if not self.tasks:
            logging.error("No tasks to export")
            return False
        
        try:
            # Generate output path if not provided
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(
                    CONFIG['paths']['default_import_path'],
                    f'imported_gantt_{timestamp}.xlsx'
                )
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Load template or create new workbook
            template_path = CONFIG['paths']['excel_template']
            if os.path.exists(template_path):
                logging.info(f"Using template: {template_path}")
                wb = openpyxl.load_workbook(template_path)
                ws = wb['Gantt Chart'] if 'Gantt Chart' in wb.sheetnames else wb.active
            else:
                logging.warning("Template not found, creating new workbook")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Gantt Chart'
                self._create_headers(ws)
            
            # Write task data
            self._write_tasks_to_excel(ws)
            
            # Save workbook
            wb.save(output_path)
            logging.info(f"Successfully exported to: {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}")
            return False
    
    def _create_headers(self, worksheet):
        """Create header row in Excel"""
        headers = [
            'Task ID', 'Task Name', 'Duration (Days)', 'Start Date', 'Finish Date',
            'Progress (%)', 'Status', 'Dependencies', 'Assigned To', 'Priority', 'Notes'
        ]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
    
    def _write_tasks_to_excel(self, worksheet):
        """Write task data to Excel worksheet"""
        start_row = 2  # Assuming headers are in row 1
        
        for idx, task in enumerate(self.tasks, start=start_row):
            worksheet.cell(row=idx, column=1, value=task['Task_ID'])
            worksheet.cell(row=idx, column=2, value=task['Task_Name'])
            worksheet.cell(row=idx, column=3, value=task['Duration_Days'])
            worksheet.cell(row=idx, column=4, value=task['Start_Date'])
            worksheet.cell(row=idx, column=5, value=task['Finish_Date'])
            worksheet.cell(row=idx, column=6, value=task['Progress'])
            worksheet.cell(row=idx, column=7, value=task['Status'])
            worksheet.cell(row=idx, column=8, value=task['Predecessors'])
            worksheet.cell(row=idx, column=9, value=task['Assigned_To'])
            worksheet.cell(row=idx, column=10, value=task['Priority'])
            worksheet.cell(row=idx, column=11, value=task['Notes'])

def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Import MS Project XML to Excel Gantt chart')
    parser.add_argument('--input', '-i', required=True, help='Path to MS Project XML file')
    parser.add_argument('--output', '-o', help='Path to output Excel file (optional)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.input):
        logging.error(f"Input file not found: {args.input}")
        return
    
    # Import MS Project file
    importer = MSProjectImporter(args.input)
    
    if importer.parse_xml():
        if importer.to_excel(args.output):
            print(f"Successfully imported {len(importer.tasks)} tasks to Excel")
        else:
            print("Failed to export to Excel")
    else:
        print("Failed to parse MS Project XML file")

if __name__ == "__main__":
    main()

